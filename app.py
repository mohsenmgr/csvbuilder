from measure_filler import Populator
from utils_cls import MyUtils
import typer
from typing_extensions import Annotated
from csv_builder import CsvBuilder
from openpyxl import load_workbook


utils = MyUtils("./")
app = typer.Typer(rich_markup_mode="rich")


@app.command(epilog="Made with 🍆 by [blue]MOSS[/blue]")
def auto(excelfile: Annotated[str, typer.Argument(help="Excel file name .xlsx")]):
    """
    Combine excel sheets together
    """

    if not excelfile.endswith(".xlsx"):
        excelfile = excelfile + ".xlsx"

    if not utils.checkIfFileExistsInPath(excelfile):
        print(f"file {excelfile} does not exist!")
        raise OSError("File does not exist")

    # Replace 'your_excel_file.xlsx' with the path to your Excel file
    workbook = load_workbook(excelfile)

    sheet_names = workbook.sheetnames
    if len(sheet_names) < 2:
        print(f"I need at least two sheets!")
        raise OSError("Give me more shit")

    builder = CsvBuilder(
        excelfile, sheet_names[0], sheet_names[1], "Hostname Edge", [])
    builder.create_output_device_list()
    builder.fill_appendix_data()
    builder.show_device_list()
    builder.combiner()
    builder.reorderData()
    builder.createCSV()


@app.command(epilog="Made with 🍆 by [blue]MOSS[/blue]")
def manual(excelfile: Annotated[str, typer.Argument(help="Excel file name .xlsx")],
           columnsfile: Annotated[str, typer.Argument(help="File containing column names .txt",
                                                      rich_help_panel="Secondary Arguments")] = 'columns_custom.txt',
           unique_name_column: Annotated[str, typer.Argument(help="Unique column, usually hostname edge",
                                                             rich_help_panel="Secondary Arguments")] = 'Hostname Edge'
           ):
    """
    Combine excel sheets, only select specified columns
    """

    if not excelfile.endswith(".xlsx"):
        excelfile = excelfile + ".xlsx"

    if not columnsfile.endswith(".txt"):
        columnsfile = columnsfile + ".txt"

    if not utils.checkIfFileExistsInPath(excelfile):
        print(f"file {columnsfile} does not exist!")
        raise OSError("File does not exist")

    if not utils.checkIfFileExistsInPath(columnsfile):
        print(f"file {columnsfile} does not exist!")
        raise OSError("File does not exist")

    populator = Populator(columnsfile)
    column_names = populator.getMeasures()

    builder = CsvBuilder(excelfile, 0, 1, unique_name_column, column_names)
    builder.create_output_device_list()
    builder.fill_appendix_data()
    builder.show_device_list()
    builder.combiner()
    builder.reorderData()
    builder.createCSV()


if __name__ == "__main__":
    app()
